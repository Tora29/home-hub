#!/usr/bin/env python3
"""Excel設計書をMarkdownテーブルに変換する。

初回実行時に .venv を自動セットアップする（Python 3 が必要）。

Usage:
    python extract_xlsx.py <path/to/file.xlsx>
"""
import subprocess, sys, os

VENV = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".venv")
PYTHON = os.path.join(VENV, "bin", "python3")

if sys.executable != PYTHON:
    if not os.path.exists(VENV):
        print("初回セットアップ: venv を作成中...", file=sys.stderr)
        subprocess.run([sys.executable, "-m", "venv", VENV], check=True)
    subprocess.run(
        [os.path.join(VENV, "bin", "pip"), "install", "-q", "openpyxl"],
        check=True,
    )
    os.execv(PYTHON, [PYTHON] + sys.argv)

import openpyxl


def sheet_to_md(ws) -> str:
    rows = [
        [str(c.value) if c.value is not None else "" for c in row]
        for row in ws.iter_rows()
        if any(c.value is not None for c in row)
    ]
    if not rows:
        return "（空のシート）"

    col = max(len(r) for r in rows)
    rows = [r + [""] * (col - len(r)) for r in rows]

    lines = (
        ["| " + " | ".join(rows[0]) + " |"]
        + ["| " + " | ".join(["---"] * col) + " |"]
        + ["| " + " | ".join(r) + " |" for r in rows[1:]]
    )
    return "\n".join(lines)


def main():
    if len(sys.argv) < 2:
        print("Usage: extract_xlsx.py <path>", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
    sections = [
        f"## シート: {name}\n\n{sheet_to_md(wb[name])}" for name in wb.sheetnames
    ]
    print("\n\n".join(sections))


main()
